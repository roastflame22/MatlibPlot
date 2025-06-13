import matplotlib.pyplot as plt
import numpy as np
import openpyxl
import os
def checkInternal(pydata, x, y):
    #grab worksheet
    DataSheet = openpyxl.load_workbook(pydata, data_only=True)

    #activate the sheet
    sheet1 = DataSheet.active

    notNone = False
    if((sheet1.cell(row=x,column=y)).internal_value != None and (sheet1.cell(row=x,column=y)).internal_value != 'none'):
        notNone = True
    return notNone
xMax = -999
yMax = -999
xMin = 999
yMin = 999
def yMaxLimit(pydata, y):
   #print("ymaxlimit set")
   #print(y)
    global yMax

    #grab worksheet
    DataSheet = openpyxl.load_workbook(pydata, data_only=True)

    #activate the sheet
    sheet1 = DataSheet.active

    if(yMax < int((sheet1.cell(row=13,column=y)).internal_value)):
        #get y Max Limit
        yMax = (sheet1.cell(row=13,column=y)).internal_value
#   print(yMax)

def yMinLimit(pydata, y):
   #print("yminlimit set")
   #print(y)
    global yMin

    #grab worksheet
    DataSheet = openpyxl.load_workbook(pydata, data_only=True)

    #activate the sheet
    sheet1 = DataSheet.active

    if(yMin > int((sheet1.cell(row=14,column=y)).internal_value)):
        #get y Min Limit
        yMin = (sheet1.cell(row=14,column=y)).internal_value
   #print(yMin)

def xMaxLimit(pydata, y):
    global xMax
    
    #grab worksheet
    DataSheet = openpyxl.load_workbook(pydata, data_only=True)

    #activate the sheet
    sheet1 = DataSheet.active

    if(xMax < (sheet1.cell(row=13, column=y).internal_value)):
        xMax = (sheet1.cell(row=13, column=y)).internal_value

def xMinLimit(pydata, y):
    global xMin
    
    #grab worksheet
    DataSheet = openpyxl.load_workbook(pydata, data_only=True)

    #activate the sheet
    sheet1 = DataSheet.active

    #get x Min Limit
    if(xMin > int((sheet1.cell(row=14,column=y)).internal_value)):
        xMin = (sheet1.cell(row=14,column=y)).internal_value

def GetAllLimits(dataList, SingleInsert, pydata, NoTests, i):
    global xMax
    global xMin
    global yMax
    global yMin

    if(SingleInsert):
        x = 3
    else:
        x = 1
    temps = np.zeros((x))
    SetXmin = np.zeros((x))
    SetXmax = np.zeros((x))
    SetYmin = np.zeros((x))
    SetYMax = np.zeros((x))
    start = 0
    y = 0
    for h in temps:
        xMax = -999
        xMin = 999
        yMax = -999
        yMin = 999
        if(checkInternal(pydata, 13, 3+start) == True and checkInternal(pydata, 14, 3+start) == True and SingleInsert == True):
            xMaxLimit(pydata, 3+start)
            xMinLimit(pydata, 3+start)
            SetXmin[y] = xMin
            SetXmax[y] = xMax
        else:
            SetXmin[y] = 23
            SetXmax[y] = 27
        if(SingleInsert):
           #print("ymax ymin true false")
           #print(checkInternal(pydata, 13, 4+start+i))
           #print(checkInternal(pydata, 14, 4+start+i))
            if(checkInternal(pydata, 13, 4+start+i) == True and checkInternal(pydata, 14, 4+start+i) == False):
                yMaxLimit(pydata, 4+start+i)
                SetYMax[y] = yMax
                SetYmin[y] = 0
            elif(checkInternal(pydata, 13, 4+start+i) == False and checkInternal(pydata, 14, 4+start+i) == True):
                yMinLimit(pydata, 4+start+i)
                SetYMax[y] = yMin * 2
                SetYmin[y] = yMin
            else:
                yMinLimit(pydata, 4+start+i)
                yMaxLimit(pydata, 4+start+i)
                SetYmin[y] = yMin
                SetYMax[y] = yMax
        else:
            if(checkInternal(pydata, 13, 3+start+i) == True and checkInternal(pydata, 14, 3+start+i) == True):
                yMinLimit(pydata, 3+start+i)
                yMaxLimit(pydata, 3+start+i)
                SetYmin[y] = yMin
                SetYMax[y] = yMax
            elif(checkInternal(pydata, 14, 3+start+i) == True and checkInternal(pydata, 13, 3+start+i) == False):
                yMinLimit(pydata, 3+start+i)
                SetYmin[y] = yMin
                SetYMax[y] = yMin * 2
            elif(checkInternal(pydata, 13, 3+start+i) == True and checkInternal(pydata, 14, 3+start+i) == False):
                yMaxLimit(pydata, 3+start+i)
                SetYmin[y] = 0
                SetYMax[y] = yMax
            
        start = start + NoTests
        y += 1
    limits = [SetXmax, SetXmin, SetYMax, SetYmin]
    return limits

def GetLimits(dataList, SingleInsert, pydata, NoTests, i):
    global xMax
    global xMin
    global yMax
    global yMin
    xMax = -999
    xMin = 999
    yMax = -999
    yMin = 999

    if(SingleInsert):
        x = 3
    else:
        x = 1
    start = 0

    temps = np.zeros((x))
    start = 0
    if(SingleInsert == True):
       #print("Single Insert: ")
        for h in temps:
            if(checkInternal(pydata, 13, 3+start) == True and checkInternal(pydata, 14, 3+start) == True):
                xMaxLimit(pydata, 3+start)
                xMinLimit(pydata, 3+start)
            if(checkInternal(pydata, 13, 4+start+i) == True and checkInternal(pydata, 14, 4+start+i) == True):
                yMinLimit(pydata, 4+start+i)
                yMaxLimit(pydata, 4+start+i)
            elif(checkInternal(pydata, 13, 4+start+i) == True and checkInternal(pydata, 14, 4+start+i) == False):
                yMaxLimit(pydata, 4+start+i)
                yMin = 0
            else:
                yMinLimit(pydata, 4+start+i)
                yMax = yMin * 2
            start = start + NoTests
       #print("yMax yMin")
       #print(yMax)
       #print(yMin)
    else:
       #print("Not single insert")
        if(checkInternal(pydata, 13, 3+i) == True and checkInternal(pydata, 14, 3+i) == True):
            yMaxLimit(pydata, 3 + i)
            yMinLimit(pydata, 3 + i)
        elif(checkInternal(pydata, 13, 3+i) == True and checkInternal(pydata, 14, 3+i) == False):
            yMaxLimit(pydata, 3 + i)
            yMin = 0
        elif(checkInternal(pydata, 13, 3+i) == False and checkInternal(pydata, 14, 3+i) == True):
            yMinLimit(pydata, 3 + i)
            yMax = yMin * 2
        else:
            yMin = 888
            yMax = 888
        xMax = 27
        xMin = 23
        ydatalist = np.zeros(len(dataList[0]))
        u = 0
        for h in ydatalist:
            ydatalist[u] = 25
            u += 1

    ExpandLimits()
    if(SingleInsert):
        xMaxxMinyMaxyMinYdata = [xMax, xMin, yMax, yMin]
    else:
        xMaxxMinyMaxyMinYdata = [xMax, xMin, yMax, yMin, ydatalist]

    return xMaxxMinyMaxyMinYdata

def ExpandLimits():
    global xMax
    global xMin
    global yMax
    global yMin
   #print("expanding Limits Pre")
   #print(xMax)
   #print(xMin)
   #print(yMax)
   #print(yMin)
    if(yMin != None and yMin != 'none'):
        if(yMin < 0 and yMax != None and yMax != 'none'):
            yMin = yMin * 1.2
            yMax = yMax * 1.2
        elif(yMin > 0 and yMax != None and yMax != 'none'):
            yMin = yMin * 0.8
            yMax = yMax * 1.2
        elif(yMin < 0):
            yMin = yMin * 1.2
            yMax = 0
        elif(yMin > 0):
            yMin = yMin * 0.8
            yMax = yMax * 2
        else:
            yMin = 0
            yMax = yMax * 1.2
        if(xMin != None, yMin != 'none'):
            if(xMin < 0):
                xMin = xMin * 1.2
                xMax = xMax * 1.2
            else:
                xMin = xMin * 0.8
                xMax = xMax * 1.2
import matplotlib.pyplot as plt
import numpy as np
import openpyxl
import DataConversion
import os
import GetLimits

def PlotTempVsTest(NoTests, SingleInsert, dataList, pydata, RunName, Path):
    #grab worksheet
    DataSheet = openpyxl.load_workbook(pydata, data_only=True)

    #activate the sheet
    sheet1 = DataSheet.active
    y = NoTests;
    if(SingleInsert):
        x = 3
        y = NoTests - 1;
        numbTest = np.zeros((y))
    else:
        x = 1
        y = NoTests;
        numbTest = np.zeros((y))
    try:
        os.mkdir(Path)
        print(f"folder '{Path}' created successfully")
    except FileExistsError:
        print(f"folder '{Path}' already exists.")
    except Exception as e:
        print(f"An error occured: {e}")
    Path = Path + "/"
    initalPath = Path


    temps = np.zeros((x))
    i = 0
    #add loop for temps here
    for c1 in numbTest:
        Path = initalPath
        xName       = "Temperature"
        for c2 in temps:
            Title       = RunName + "_Temperature_Vs_"
            yName       = ""
            if(SingleInsert):
                yName = (sheet1.cell(row=9, column=4+i)).internal_value
            else:
                yName = (sheet1.cell(row=9, column=3+i)).internal_value
            Title = Title + yName

        xyMaxMinYdata = GetLimits.GetLimits(dataList, SingleInsert, pydata, NoTests, i)

        #add plot function here
        print("XXXXXxxxxxx start of New plot Function xxxxXXXXX")
        print(i + 1)
       #print("newset of limits are")
       #print(xyMaxMinYdata[0])
       #print(xyMaxMinYdata[1])
       #print(xyMaxMinYdata[2])
       #print(xyMaxMinYdata[3])
       #print(xMaxLimit)
       #print(xMinLimit)
       #print(yMaxLimit)
       #print(yMinLimit)

        limits = GetLimits.GetAllLimits(dataList, SingleInsert, pydata, NoTests, i)
        xLineYmaxYmin = DataConversion.GetLimitPlot(dataList, SingleInsert, pydata, NoTests, i, limits)
        

            

        plt.style.use('_mpl-gallery')
        with plt.style.context('dark_background'):
            plt.figure(figsize=(10,6))
            plt.subplots_adjust(top=0.95, bottom=0.1, left=0.1, right=0.95)
            Titleformat = Title.replace(" ", "")
            Titleformat = Titleformat.replace(".", "")
            Titleformat = Titleformat.replace("/","")
            Path = Path + Titleformat
           #print(Path + ".png")
            if(os.path.exists(Path + ".png") == False):
                Path = Path + ".png"
               #print("3V image made")
               #print(Path)
                
            else:
                Path = Path + "3_6V.png"
               #print("3.6V image made")
               #print(Path)

            #add CPK
            cpk = DataConversion.CPK(NoTests, SingleInsert, dataList, pydata, i)
            if(SingleInsert):
                if(cpk[0] == "inf"):
                    Title = Title + " 25 C CPK: " + "inf" + " -55C CPK: " + str(round(cpk[1], 2)) + " 85C CPK: " + str(round(cpk[2], 2))
                elif(cpk[1] == "inf"):
                    Title = Title + " 25 C CPK: " + str(round(cpk[0], 2)) + " -55C CPK: " + "inf" + " 85C CPK: " + str(round(cpk[2], 2))
                elif(cpk[2] == "inf"):
                    Title = Title + " 25 C CPK: " + str(round(cpk[0], 2)) + " -55C CPK: " + str(round(cpk[1], 2)) + " 85C CPK: " + "inf"
                else:
                    Title = Title + " 25 C CPK: " + str(round(cpk[0], 2)) + " -55C CPK: " + str(round(cpk[1], 2)) + " 85C CPK: " + str(round(cpk[2], 2))
            else:
                if(cpk[0] == "inf"):
                    Title = Title + " 25 C CPK: " + "inf"
                else:
                    Title = Title + " 25 C CPK: " + str(round(cpk[0], 2))
            print(Title)
           #print("xyMaxMinYdata")
           #print(xyMaxMinYdata)
            plt.xlim(xyMaxMinYdata[1], xyMaxMinYdata[0])
            plt.ylim(xyMaxMinYdata[3], xyMaxMinYdata[2])
            plt.xlabel(xName, fontsize = 12)  #Title, text size
            plt.ylabel(yName, fontsize = 12)    #Title, text size
            plt.title(Title, fontsize = 15)
            if(SingleInsert):
               #print(dataList[start])
               #print(dataList[start+i])
               #print("Xline ymax ymin")
               #print(xLineYmaxYmin[0])
               #print(yMax)
               #print(yMin)
                plt.plot(dataList[0], dataList[i + 1], 'gx', markeredgewidth=2)
                plt.plot(dataList[NoTests], dataList[1+i+NoTests], 'bx', markeredgewidth=2)
                plt.plot(dataList[NoTests*2], dataList[1+i+NoTests*2], 'rx', markeredgewidth=2)
                plt.plot(xLineYmaxYmin[0], xLineYmaxYmin[1], color='darkred', linestyle='--', markeredgewidth=2)
                plt.plot(xLineYmaxYmin[0], xLineYmaxYmin[2], color='aqua', linestyle='--', markeredgewidth=2)
            else:
               #print(i)
               #print(xxLineYmaxYmin[1]MinYdata[4])
               #print(dataList[i])
                plt.plot(xyMaxMinYdata[4], dataList[i], 'gx', markeredgewidth=2)
                plt.plot(xLineYmaxYmin[0], xLineYmaxYmin[1], color='darkred', linestyle='--', markeredgewidth=2)
                plt.plot(xLineYmaxYmin[0], xLineYmaxYmin[2], color='aqua', linestyle='--', markeredgewidth=2)
            plt.savefig(Path, dpi=400)
        i = i + 1
        
#   plt.style.use('_mpl-gallery')
#   with plt.style.context('dark_background'):
#       plt.figure(figsize=(10, 6))
#       plt.subplots_adjust(top=0.95, bottom=0.1, left=0.07, right=.95)
#       plt.ylim(DataConversion.yMinLimit * 0.8, DataConversion.yMaxLimit * 1.2)
#       plt.xlim(DataConversion.xMinLimit * 1.2, DataConversion.xMaxLimit * 1.2)
#       plt.xlabel(DataConversion.xName, fontsize = 12)  #Title, text size
#       plt.ylabel(DataConversion.yName, fontsize = 12)    #Title, text size
#       plt.plot(RealTemp25C, RealVf25C, 'gx', markeredgewidth=2)
#       plt.plot(RealTempN55C, RealVfN55C, 'bx', markeredgewidth=2)
#       plt.plot(RealTemp85C, RealVf85C, 'rx', markeredgewidth=2)
            


#       plt.style.use('_mpl-gallery')
#       with plt.style.context('dark_background'):
#           plt.figure(figsize=(10, 6))
#           plt.subplots_adjust(top=0.95, bottom=0.1, left=0.07, right=.95)
#           plt.ylim(DataConversion.yMinLimit * 0.8, DataConversion.yMaxLimit * 1.2)
#           plt.xlim(DataConversion.xMinLimit * 1.2, DataConversion.xMaxLimit * 1.2)
#           plt.xlabel(DataConversion.xName, fontsize = 12)  #Title, text size
#           plt.ylabel(DataConversion.yName, fontsize = 12)    #Title, text size
#           plt.plot(RealTemp25C, RealVf25C, 'gx', markeredgewidth=2)
#           plt.plot(RealTempN55C, RealVfN55C, 'bx', markeredgewidth=2)
#           plt.plot(RealTemp85C, RealVf85C, 'rx', markeredgewidth=2)

       #xMaxLimit   = 000
       #xMinLimit   = 999
       #yMaxLimit   = 000
       #yMinLimit   = 999

#           if((sheet1.cell(row=13,column=3+start)).internal_value != None and (sheet1.cell(row=14,column=3+start)).internal_value != None and SingleInsert == True):
#               if(xMaxLimit == 0):
#                   xMaxLimit = (sheet1.cell(row=13,column=3+start)).internal_value
#               elif(xMaxLimit < (sheet1.cell(row=13,column=3+start).internal_value)):
#                   xMaxLimit = (sheet1.cell(row=13,column=3+start)).internal_value
#               #get x Min Limit
#               if(xMaxLimit == 999):
#                   xMinLimit = (sheet1.cell(row=14,column=3+start)).internal_value
#               elif(xMinLimit > int((sheet1.cell(row=14,column=3+start)).internal_value)):
#                   xMinLimit = (sheet1.cell(row=14,column=3+start)).internal_value
#               #get title for graph
#               yName = (sheet1.cell(row=9, column=4+i)).internal_value
#               Title = Title + yName
#               #print(Title)
#           else:
#               print("Not single insert")
#               xMaxLimit = 27
#               xMinLimit = 23
#               yName = (sheet1.cell(row=9, column=4+i)).internal_value
#               Title = Title + yName
#               ydatalist = np.zeros(len(dataList[0]))
#               u = 0
#               for h in ydatalist:
#                   ydatalist[u] = 25
#                   u += 1
#           print((sheet1.cell(row=13,column=4+start+i)).internal_value)
#           print((sheet1.cell(row=14,column=4+start+i)).internal_value)
#           if((sheet1.cell(row=13,column=4+start+i)).internal_value != None and (sheet1.cell(row=14,column=4+start+i)).internal_value != None):
#               if(yMaxLimit < int((sheet1.cell(row=13,column=4+start+i)).internal_value)):
#                   #get y Max Limit
#                   yMaxLimit = (sheet1.cell(row=13,column=4+start+i)).internal_value
#               if(yMinLimit > int((sheet1.cell(row=14,column=4+start+i)).internal_value)):
#                   #get y Min Limit
#                   yMinLimit = (sheet1.cell(row=14,column=4+start+i)).internal_value
#           if(yMinLimit != None and yMinLimit != 'none'):
#               if(yMinLimit < 0 and yMaxLimit != None and yMaxLimit != 'none'):
#                   plt.ylim(yMinLimit * 1.2, yMaxLimit * 1.2)
#               elif(yMinLimit > 0 and yMaxLimit != None and yMaxLimit != 'none'):
#                   plt.ylim(yMinLimit * 0.8, yMaxLimit * 1.2)
#               elif(yMinLimit < 0):
#                   plt.ylim(yMinLimit * 1.2, 0)
#               elif(yMinLimit > 0):
#                   plt.ylim(yMinLimit * 0.8, yMinLimit * 2)
#           else:
#               plt.ylim(0, yMaxLimit *1.2)
#           if(xMinLimit != None):
#               if(xMinLimit < 0):
#                   plt.xlim(xMinLimit * 1.2, xMaxLimit * 1.2)
#               else:
#                   plt.xlim(xMinLimit * 0.8, xMaxLimit * 1.2)
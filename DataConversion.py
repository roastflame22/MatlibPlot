import matplotlib.pyplot as plt
import numpy as np
import openpyxl
import statistics

import openpyxl.workbook
import openfiles

SingleInsert_test = 0
xMaxLimit   = 000
xMinLimit   = 999
yMaxLimit   = 000
yMinLimit   = 999
xName       = "xxx"
yName       = "xxx"
Title       = "xxx"

def GethasTemp(pydata):
    hasTemp = False
    #grab worksheet
    DataSheet = openpyxl.load_workbook(pydata, data_only=True)

    #activate the sheet
    sheet1 = DataSheet.active

    i = 1
    for tests in pydata: 
        if((sheet1.cell(row=9,column=3+i)).internal_value != None):
            title = (sheet1.cell(row=9,column=3+i)).internal_value
            #print("title")
            #print(title)
            if(title == "Temperature"):
                #print("first Temp")
                hasTemp = True
            i += 1
    return hasTemp
def GetTemps(NoTests, datalist, pydata, i):
    #print("pydata in get temps")
    #print(pydata)
    #grab worksheet
    DataSheet = openpyxl.load_workbook(pydata, data_only=True)

    #activate the sheet
    sheet1 = DataSheet.active

    Temps = ['empty', 'empty', 'empty']
    Temps[0] = str((sheet1.cell(row=13, column=3)).internal_value - 2)
    Temps[1] = str((sheet1.cell(row=13, column=3+NoTests)).internal_value - 2)
    Temps[2] = str((sheet1.cell(row=13, column=3+(NoTests * 2))).internal_value - 2)
        
    #print("Temp List")
    #print(Temps[0])
    #print(Temps[1])
    #print(Temps[2])
    return Temps
def GetTester(pydata):
    TesterString = pydata.split("T")
    Tester = "T" + TesterString[1]
    #print("tester")
    #print(Tester)
    return Tester
def GetSingleInsert(pydata, filenumber):
    SingleInsert = False
    Notemp = 0

    #grab worksheet
    DataSheet = openpyxl.load_workbook(pydata[filenumber], data_only=True)

    #activate the sheet
    sheet1 = DataSheet.active

    i = 1
    for tests in pydata[filenumber]: 
        if((sheet1.cell(row=9,column=2+i)).internal_value != None):
            title = (sheet1.cell(row=9,column=2+i)).internal_value
            #print("title")
            #print(title)
            if(title == "Temperature"):
                print("first Temp")
                Notemp = Notemp + 1
            i += 1
    if(Notemp > 0):
        SingleInsert = True
    return SingleInsert

def GetNoTests(pydata):
    x = 160
    i = 1
    firstTemp = True
    hasTemp = False
    datalist = np.zeros(x)
    NoTests = np.zeros(len(pydata))
    currtest = 0
    for data in pydata:
        #grab worksheet
        DataSheet = openpyxl.load_workbook(pydata[currtest], data_only=True)

        #activate the sheet
        sheet1 = DataSheet.active
        for tests in datalist: 
            if((sheet1.cell(row=9,column=3+i)).internal_value != None):
                title = (sheet1.cell(row=9,column=3+i)).internal_value
                #print("title")
                #print(title)
                if(title == "Temperature" and firstTemp == True):
                    #print("first Temp")
                    firstTemp = False
                    hasTemp = True
                    NoTests[currtest] = i
               #    #print("Notests")
               #    #print(NoTests)
                else:
                    i += 1
        if(hasTemp == False):
            NoTests[currtest] = i
        firstTemp = True
        currtest = currtest + 1
        i = 1
    return NoTests

def GetTestNames(pydata, data, SingleInsert):
    DataSheet = openpyxl.load_workbook(pydata, data_only=True)
    sheet1 = DataSheet.active
    #print("GetTestNames")
    #print("TestNames length = len(data) =" + str(len(data) / 3 - 1))
    
    if(SingleInsert):
        TestNames = [" "] * int((len(data)/3 - 1))
    else:
        TestNames = [" "] * int((len(data)))
    i = 0
    for x in data:
        #print("i")
        #print(i)
        if(SingleInsert):
            if(i < len(data) / 3 - 1):
                TestNames[i] = (sheet1.cell(row=9, column=4+i)).internal_value
                #print(TestNames[i - 3])
            else:
                return TestNames
        else:
            if(i < len(data) & SingleInsert != True):
                TestNames[i] = (sheet1.cell(row=9, column=3+i)).internal_value
            else:
                return TestNames
        i += 1
    return TestNames


def Conversion(NoTests, SingleInsert, pydata):
    #grab worksheet
    DataSheet = openpyxl.load_workbook(pydata, data_only=True)

    #activate the sheet
    sheet1 = DataSheet.active
    isTemp = False;
    if(SingleInsert):
        SingleInsert_test = NoTests * 3
        x = SingleInsert_test + 1
        y = 256
        a = SingleInsert_test + 1
        b = 256
       #x, y = SingleInsert_test, len(pydata)
       #a, b = SingleInsert_test, len(pydata)- 10
    else:
        x, y = NoTests, len(pydata)
        a, b = NoTests, len(pydata)- 10
#    x, y = NoTests, DataLength
    #print("x, y , a, b")
    #print(x)
    #print(y)
    #print(a)
    #print(b)
    array2D = np.zeros((x,y))
    datalist = np.zeros((a,b))

    cols = 3
    for col in array2D:
        rows = 9
        i = 0
        for row in col:   
            if(i >= 10):
                datalist[cols-3][i-10] = (sheet1.cell(row=rows,column=cols)).internal_value
                #print(array2D[cols-3][rows-9])
            i = i + 1
            rows = rows + 1
        cols = cols + 1
    
    #print("datalist")
    #print(datalist)
    shortList = shortenDatalist(datalist, pydata)
    #print("shortList")
    #print(shortList)
    return shortList

def shortenDatalist(datalist, pydata):
    #grab worksheet
    DataSheet = openpyxl.load_workbook(pydata, data_only=True)

    #activate the sheet
    sheet1 = DataSheet.active

    newCol = 0
    newRow = 0

    rows = 1
    maxrow = 1
    cols = 1
    for col in datalist:
        for row in col:
            if((sheet1.cell(row=18+rows,column=2 + cols)).internal_value != None):
                newRow += 1
                rows += 1
        rows = 1
        if((sheet1.cell(row=18+rows,column=2+cols)).internal_value != None):
            newCol += 1
            cols += 1
        if(maxrow < newRow):
            maxrow = newRow
        newRow = 0

    rowCol = np.zeros((newCol, maxrow))
    rows = 0                
    cols = 0

    for col in rowCol:
        for row in col:
            rowCol[cols, rows] = datalist[cols, rows]
            rows += 1
        cols += 1
        rows = 0
    #print('newcol')
    #print(newCol)
    #print('newRow')
    #print(maxrow)
    return rowCol

def sigma_numpy(lower_bound, upper_bound, expression):
    numbers = np.arange(lower_bound, upper_bound + 1)
    return np.sum(expression(numbers))

def GageRnR(Tester, data, pydata):
    print("entering GageRnR data conversion")
    #print(data)
   ##create the sheet
    dataSheet = openpyxl.Workbook()

    #activate the sheet
    sheet1 = dataSheet.active

    basePath = openfiles.askdirectory() + "/"
    paths = [""] * len(data[0])
    print("data[0][0]")
    print(data[0][0])
    DUTNumber = len(data[0][0])

    TesterCounter   = 0
    Cols            = 0
    Rows            = 0
    DUT             = 0
    i = 0
    for Tests in data[0]:
        for datapoint in data[0][0]:
            if((sheet1.cell(row=9, column=3+i)).internal_value != None):
                yName = (sheet1.cell(row=9, column=3+i)).internal_value
            else:
                yName = "wasnull"
            path = basePath + yName + ".csv"
            paths[i] = path
            GaugeFile = open(path)
            GaugeFile.write(Tester[TesterCounter] + ";" + Rows + ";" + data[Tester[TesterCounter],Cols,Rows * DUT])
            if(DUT > DUTNumber * DUTNumber - DUTNumber - 1):
                DUT = 0
                Rows += 1
            else:
                DUT = DUT + DUTNumber
        TesterCounter += 1
        Rows = len(data[0][0])
        for datapoint in data[0][0]:
            GaugeFile.write(Tester[TesterCounter] + ";" + Rows + ";" + data[Tester[TesterCounter],Cols,(Rows / 2) * DUT ])
            if(DUT > DUTNumber * DUTNumber - DUTNumber - 1):
                DUT = 0
                Rows += 1
            else:
                DUT = DUT + DUTNumber
        Cols += 1


    return paths

def CPK(NoTests, SingleInsert, dataList, pydata, i):
    print("CPK block")
    print("NoTests")
    print(NoTests)
    #grab worksheet
    DataSheet = openpyxl.load_workbook(pydata, data_only=True)

    #activate the sheet
    sheet1 = DataSheet.active

    if(SingleInsert):
        x = 3
    else:
        x = 1

    numbTemp = np.zeros((x))

    numbTest = np.zeros((NoTests - 1))
    temps = np.zeros((x))
    Cpk = np.zeros(( x))
    start = 0
    #add loop for temps here

    y = 0
    yMaxLimit   = 000
    yMinLimit   = 999
    xName       = "Temperature"
    for c2 in temps:

        if(SingleInsert):
            if((sheet1.cell(row=13,column=4+start+i)).internal_value != None):
                #get y Max Limit
                yMaxLimit = (sheet1.cell(row=13,column=4+start+i)).internal_value
            if((sheet1.cell(row=14,column=4+start+i)).internal_value != None): 
                #get y Min Limit
                yMinLimit = (sheet1.cell(row=14,column=4+start+i)).internal_value
        else:
            if((sheet1.cell(row=13,column=3+start+i)).internal_value != None):
                #get y Max Limit
                yMaxLimit = (sheet1.cell(row=13,column=3+start+i)).internal_value
            if((sheet1.cell(row=14,column=3+start+i)).internal_value != None): 
                #get y Min Limit
                yMinLimit = (sheet1.cell(row=14,column=3+start+i)).internal_value

        print("datalist for sigma for calculations")
        if(SingleInsert):
            #print(dataList[start + i + 1])
            sigma = np.std(dataList[start + i + 1])
            mean  = np.mean(dataList[start + i + 1])
        else:
            #print(dataList[start + i])
            sigma = np.std(dataList[start + i])
            mean  = np.mean(dataList[start + i])
        #print("sigma and mean")
        #print(sigma)
        #print(mean)
        #print("ymax ymin")
        #print(yMaxLimit)
        #print(yMinLimit)
        if(SingleInsert):
            if((sheet1.cell(row=14,column=4+start+i)).internal_value != None and (sheet1.cell(row=13,column=4+start+i)).internal_value != None): 
                if(yMinLimit != "none"):
                    #print("Cpl")
                    Cpl = (mean - yMinLimit) / (3 * sigma)
                    #print(Cpl)
                else:
                    Cpl = 9999
                if(yMaxLimit != "none"):
                    #print("Cpu")
                    Cpu = (yMaxLimit - mean) / (3 * sigma)
                    #print(Cpu)
                else:
                    Cpu = 9999
                Cpk[y] = min(abs(Cpu),abs(Cpl))
                start = start + NoTests
                y += 1
        else:
            if(yMinLimit != "none"):
                #print("Cpl")
                Cpl = (mean - yMinLimit) / (3 * sigma)
                #print(Cpl)
            else:
                Cpl = (mean - 0) / (3 * sigma)
            if(yMaxLimit != "none"):
                #print("Cpu")
                Cpu = (yMaxLimit - mean) / (3 * sigma)
                #print(Cpu)
            else:
                Cpu = (yMinLimit * 2 - mean) / (3 * sigma)
            Cpk[y] = min(abs(Cpu),abs(Cpl))
            start = start + NoTests
            y += 1

    #print("Cpk")
    #print(Cpk)
    return Cpk




def ConversionXL(DataList):
    #global Vars for title and limits
    global xMaxLimit
    global xMinLimit
    global yMaxLimit
    global yMinLimit
    global xName
    global yName
    global Title

    RealTemp25C = [0]
    i = 0
    isTemp = False
    for c1, c2 in DataList:
        if(i == 0):
            Title = c1.internal_value
            if(Title == "Temperature"):
                isTemp = True
                xName = c1.internal_value
                #print("xName")
                #print(xName)
            else:
                yName = c1.internal_value
                #print("yName")
                #print(yName)
        elif(i == 3):
            Unit = c1.internal_value
        elif(i == 4):
            if(isTemp == True):
                templimit = c1.internal_value
                xMaxLimit = CheckMaxLimit(templimit, xMaxLimit)
            else:
                templimit = c1.internal_value
                yMaxLimit = CheckMaxLimit(templimit, yMaxLimit)
        elif(i == 5):
            if(isTemp == True):
                templimit = c1.internal_value
                xMinLimit = CheckMinLimit(templimit, xMinLimit)
            else:
                templimit = c1.internal_value
                yMinLimit = CheckMinLimit(templimit, yMinLimit)
        if(i > 10):
            if(RealTemp25C == 0):
                RealTemp25C.insert(0, c1.internal_value)  
            else:
                #print("test")
                RealTemp25C.append(c1.internal_value)
                #print(c1.internal_value)
        i = i + 1
        #print(i)
        #print(c1.internal_value)
    return RealTemp25C
def CheckMinLimit(tempLimit, oldLimit):
    if(tempLimit < oldLimit):
        newLimit = tempLimit
    else:
        newLimit = oldLimit
    return newLimit
        
def CheckMaxLimit(tempLimit, oldLimit):
    if(tempLimit > oldLimit):
        newLimit = tempLimit
    else:
        newLimit = oldLimit
    return newLimit

def GetLimitPlot(dataList, SingleInsert, pydata, NoTests, i, limits):
    #print("limits")
    #print(limits)
    h = 0
    xMax = limits[0]
    xMin = limits[1]
    xLine = np.zeros((40))
    yMax = np.zeros((40))
    yMin = np.zeros((40))
    #print("xMax[2]: ")
    #print(xMax)
    #print("xMin[1]: ")
    #print(xMin)

    if(SingleInsert):
        stepSize = (xMax[2] - xMin[1]) / 40
        print("stepsize")
        print(stepSize)
        print("starting point")
        print(xMin[1] - stepSize * 1.2)

        for data in xLine:
            if(xLine[0] == 0):
                xLine[0] = int(xMin[1] * 1.2)
                yMax[h] = limits[2][1]
                yMin[h] = limits[3][1]
            elif(xLine[h-1] < -20):
                xLine[h] = xLine[h-1] + stepSize * 1.2
                yMax[h] = limits[2][1]
                yMin[h] = limits[3][1]
            elif(xLine[h-1] > -20 and xLine[h-1] < 40):
                xLine[h] = xLine[h-1] + stepSize * 1.5
                yMax[h] = limits[2][0]
                yMin[h] = limits[3][0]
            else:
                xLine[h] = xLine[h-1] + stepSize * 2
                yMax[h] = limits[2][2]
                yMin[h] = limits[3][2]
            #print("xLine" + str(xLine[h]))
            h += 1
    else:
        stepSize = (27 - 23) / len(dataList[0])
        #print("stepsize")
        #print(stepSize)
        #print("starting point")
        #print(xMin[0] - stepSize * 1.2)

        for data in dataList[0]:
            if(xLine[0] == 0):
                #print("xMin")
                #print(xMin[0])
                xLine[0] = xMin[0] - stepSize*1.2
            else:
                xLine[h] = xLine[h-1] + stepSize * 1.2
                yMax[h] = limits[2][0]
                yMin[h] = limits[3][0]
            h += 1
    xLineYmaxYmin = [xLine, yMax, yMin]
    return xLineYmaxYmin